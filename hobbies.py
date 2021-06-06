"""
Advance Python Project
Author-Nikhil Pothu
Date-04-06-2021
"""
import openpyxl


def hobbies_data(INC,excel_name):
    """ Data extraction and storing in new file"""
    excel_document = openpyxl.load_workbook('main.xlsx')
    w_s = excel_document['Hobbies']
    print("Available Data: ")
    for col in range(1, w_s.max_column + 1):
        heading = w_s.cell(row=1, column=col).value
        print(heading)
    print()
    print("Available Ps-no")
    for r_o in range(2, w_s.max_row + 1):
        ps_no = w_s.cell(row=r_o, column=1).value
        print(ps_no)
    print()
    try:
        wb1 = openpyxl.load_workbook(excel_name)
    except FileNotFoundError as err:
        print(err)
    ws1 = wb1.active
    print("Please Select Ps-No")
    try:
        ps_no1 = input()
        for p_s in range(2, w_s.max_row + 1):
            ps_no = str(w_s.cell(row=p_s, column=1).value)
            if ps_no == ps_no1:
                for abc in range(1, w_s.max_column + 1):
                    data = w_s.cell(row=p_s, column=abc)
                    data1 = w_s.cell(row=1, column=abc)
                    ws1.cell(row=INC, column=abc).value = data1.value
                    ws1.cell(row=INC + 1, column=abc).value = data.value
                wb1.save(excel_name)
                return 1
                break
    except OSError as err:
        print("OS error: {0}".format(err))
        return 1
