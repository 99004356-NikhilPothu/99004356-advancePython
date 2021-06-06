"""
Advance Python Project
Author-Nikhil Pothu
Date-04-06-2021
"""
from profile import profile_data
import openpyxl
from academics import academics_data
from skills import skills_data
from hobbies import hobbies_data
from cities import cities_data


class SheetCheck:
    """Checks the sheet"""
    @staticmethod
    def sheet_checking(sheet_name):
        """Checks the sheet"""
        if sheet_name == 1:
            profile_data(INC, excel_name)
        elif sheet_name == 2:

            academics_data(INC, excel_name)
        elif sheet_name == 3:

            skills_data(INC, excel_name)
        elif sheet_name == 4:
            hobbies_data(INC, excel_name)
        elif sheet_name == 5:
            cities_data(INC, excel_name)
class Main:
    """ Main class to call specific sheet"""

    @staticmethod
    def console_data():
        """ Main method to call or specific data"""
        print("Available sheets:")

        sheet_names = excel_document.sheetnames
        i=0
        for sheets in sheet_names:
            i +=1
            print(str(i)+"."+sheets)
        print("please enter the sheet number:")
        try:
            sheet_name = int(input())
            s_1 = SheetCheck()
            s_1.sheet_checking(sheet_name)
            return 1
        except OSError as err:
            print("OS error: {0}".format(err))
            return 1
        except ValueError:
            print("Oops!  That was no valid number.  Try again...")
            return 0


if __name__ == "__main__":
    print("Welcome")
    print("Please Enter Your Name:")
    name = input()
    wb = openpyxl.Workbook()
    excel_name = name + ".xlsx"
    wb.save(excel_name)
    INC = 1
    while INC:
        print("Choose Option:")
        print("Do you want to request data:")
        print("1.Yes\n2.No")

        try:
            choice = int(input())
            if choice == 1:
                INC = INC + 2
                try:
                    excel_document = openpyxl.load_workbook('main.xlsx')
                except FileNotFoundError as err:
                    print(err)
                c = Main()
                c.console_data()
            else:
                break
        except ValueError:
            print("Oops!  That was no valid number.  Try again...")
